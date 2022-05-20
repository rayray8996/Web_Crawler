from openpyxl import load_workbook
import re

RTX_3090_Ti= []
RTX_3080_Ti=[]
RTX_3070_Ti=[]
RTX_3080=[]
RTX_3060 = []
RTX_3050 = []

RTX_A5500=[]
RTX_A5000=[]

RX_6900_XT=[]
RX_6800_XT=[]

Radeon_RX_6500=[]

change = ''
Another =[]
Cheap = 0
wb = load_workbook('test.xlsx')
sheet = wb['first']
cellRange = sheet['A1':'A10']

class start():
    def RTX_number (change):
        
        if "3050" in change:
            RTX_3050.append(change)

        elif "3060" in change:
            RTX_3060.append(change)
           
        elif "3080" in change:
            RTX_3080.append(change)

            

    #--------------------------------------------
    def search_RTX(self):
        for row in cellRange:
            for c in row:
                print(c.value)
                change = c.value
                if 'RTX' in c.value:
                    start.RTX_number(change)
                    print("pass")
                else:
                    Another.append(change)
                    print("Fail")
            print('\n')
            for i in range(len(RTX_3050)):
                sheet.cell(row = i+1,column = 5,value = RTX_3050[i])

            for i in range(len(RTX_3060)):
                sheet.cell(row = i+1,column = 2,value = RTX_3060[i])
            wb.save('test.xlsx')
    #--------------------------------------------
    def typestyle(self,RTX_3050_Price):
        self.RTX_3050_Price = RTX_3050_Price
        n = len(self.RTX_3050_Price)
        self.Cheap =  self.RTX_3050_Price[0]
        print(n)
        if n > 1:
           for i in range(n):        
               if self.Cheap > self.RTX_3050_Price[i]:  
                   self.Cheap = self.RTX_3050_Price[i]
                   
        sheet.cell(row = 1,column = 8 ,value = '最便宜'+self.Cheap)
        wb.save('test.xlsx')
            
        print(self.Cheap)
    #--------------------------------------------               
    def Go_Search(self):

        self.RTX_3050_Price =[]
        Start = RTX_3050[0].find('金額')
        End = len(RTX_3050[0]) 
        Search = (RTX_3050[0][Start:End])
        
        
        n = len(RTX_3050)
        if len(RTX_3050) > 1:
            for i in range(n):
                Start = RTX_3050[i].find('金額')
                End = len(RTX_3050[i]) 
                Search = (RTX_3050[i][Start:End])
                self.RTX_3050_Price.append(Search)
        #print(self.RTX_3050_Price)
        
        data = start.typestyle(self,self.RTX_3050_Price)
        

        
    
    
if __name__ == '__main__':
    a = start()
    c = a.search_RTX()
    b =a.Go_Search()
    
  


     