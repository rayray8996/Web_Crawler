import openpyxl
import Pchome_Search as PS
import Shopee_Search as SS
import assignment
import numpy as np

PS_Product = [] 
SS_Product = []
PS_Price = [] 
SS_Price =[]

step = 0

workbook =openpyxl.Workbook()
sheet = workbook.create_sheet('first',0)
#sheet.cell(row=1, column=1, value="python")
#sheet.cell(row=1, column=2, value="python")

sheet.cell(row=1, column=1, value=0)

            
#--------------------------------------------
if __name__ == '__main__':
    #keyword = input("請輸入查尋項目:")
    keyword = 'RTX'
    Pchome_Spider = PS.PchomeSpider()
    PS_Product = Pchome_Spider.search_products(keyword)
    for i in range(10):    
        print(PS_Product[i]['name'],PS_Product[i]['price'],i)
        price = str(PS_Product[i]['price'])
        PS_Price.append(price)
        string = PS_Product[i]['name'] + PS_Price[i] 
#--------------------------------------------      
    print("\n")
      
    Shopee_Spider = SS.ShopeeSpider()
    SS_Product = Shopee_Spider.search_products(keyword)
    for i in range(10):
        price = str(SS_Product [i]['item_basic']['price'])
        Change_price = price.removesuffix('00000')
        SS_Price.append(Change_price)
        print(SS_Product [i]['item_basic']['name'],Change_price,i)
 
#--------------------------------------------       
    for i in range(10):
        sheet.cell(row=i+1, column=1, value=PS_Product[i]['name']+'金額'+ PS_Price[i] )
        
    for i in range(10):
        sheet.cell(row=i+12, column=1, value=SS_Product[i]['item_basic']['name']+'金額'+ SS_Price[i] )

#--------------------------------------------
workbook.save('test.xlsx')   